# -*- coding: utf-8 -*-
import os
import sys
base_dir = os.path.dirname(os.path.abspath(__file__))
if base_dir not in sys.path:
    sys.path.insert(0,base_dir)
import shelve
import copy
import pandas as pd
import numpy as np
import xlrd
import unicodedata
import openpyxl


def is_number(s):
    '''最后更新日期:2023.01.06
        判断字符串是不是数字
    '''
    try:
        if s=='' or s=='nan' or '_' in s or s[-1]=='.':
            return False
        float(s)
        return True
    except ValueError:
        pass
    try:
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False

def return_number(s,reserved_decimal='off'):
    if not is_number(s) or s == 'inf' or s=='-inf':
        return s
    else:
        if ('.' in s or 'e' in s )and set(s.split('.')[-1])!={'0'}:
            if reserved_decimal!='off':
                return round(float(s),reserved_decimal)
            return float(s)
        else:
            if '.' in s and set(s.split('.')[-1])=={'0'}:
                s=s[:len(s)-len(s.split('.')[-1])-1]
            try:
                return int(s)
            except:
                return s

def read_xlsx(name, sheet_num=0, turn_float=True,reserved_decimal='off'):
    workbook = xlrd.open_workbook(name)  # 打开工作簿
    data_sheet = workbook.sheets()[sheet_num]  # 通过索引获取
    rowNum = data_sheet.nrows  # sheet行数
    # 获取所有单元格的内容
    datalist = []
    for i in range(1,rowNum):#不要辩题行
        datalist.append(data_sheet.row_values(i))

    for i in range(len(datalist)):
        for j in range(len(datalist[i])):
            if isinstance(datalist[i][j],float):
                if '.' in str(datalist[i][j]) and str(datalist[i][j]).split('.')[1]=='0':
                    datalist[i][j]=int(datalist[i][j])
            if turn_float:
                datalist[i][j]=return_number(str(datalist[i][j]),reserved_decimal)
    return datalist

def write_xlsx(name, datalist= [[]],turn_float=True):
    # 设置文件 mingc
    workbook = openpyxl.Workbook()
    if isinstance(datalist,pd.core.frame.DataFrame):
        datalist=np.array(datalist).tolist()##panda转list
    # # 创建一张新表
    sheet = workbook.create_sheet('demo',index=0)
    for row in range(len(datalist)):
        for column in range(len(datalist[row])):
            if turn_float:
                tem=return_number(str(datalist[row][column]))
            else:
                tem=str(datalist[row][column])
            sheet.cell(row = row+1, column = column+1).value = tem
    if not os.path.exists(os.path.dirname(name)):# 如果不存在则创建目录
        os.makedirs(os.path.dirname(name))
    workbook.save(name)
    print("创建" + name + '成功')


def inversion_xlsx(filename_clf,filename_xlsx,fea_set=None,behind_name='_inv'):
    '''根据拟合器反演xlsx数据,没有行'''
    save = shelve.open(filename_clf[:-4])
    clf = save['clf']
    ss = save['ss']

    data=np.array(read_xlsx(filename_xlsx)[0:])
    data1=copy.deepcopy(data)
    if fea_set=='auto':
        fea_set=save['comb_x']
    else:
        if len(save['comb_x'])!=len(data1[0]):
            print(len(save['comb_x']),'(反演文件)','!=',len(data1[0]),'(文件),全部特征数量不对')
            sys.exit()
        fea_set=[x for x in range(len(data1[0]))]
    #print('特征子集：',fea_set)
    data1=ss.transform(data1[:,fea_set])
    y_clf=clf.predict(data1)
    y_clf = y_clf.reshape(-1, 1)
    result=np.concatenate((y_clf, data), axis=1)

    title = [0]
    for i in range(len(data[0])):
        title.append(i)
    title= np.array(title).reshape(-1,len(title))
    result=np.concatenate((title, result), axis=0)
    filename=filename_xlsx[:-5]+behind_name+'.xlsx'
    write_xlsx(filename,result)

if __name__ == '__main__':
    filename_clf = os.path.join(base_dir, 'SVR_StandardScaler_RRMSE_2026-05-17_23-17-32.666985.dat')
    filename_xlsx = os.path.join(base_dir, 'y_test.xlsx')
    inversion_xlsx(filename_clf,filename_xlsx)