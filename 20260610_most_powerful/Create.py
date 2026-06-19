# -*- coding: utf-8 -*-
import os
import sys
base_dir = os.path.dirname(os.path.abspath(__file__))
if base_dir not in sys.path:
    sys.path.insert(0,base_dir)
import shelve
import copy
import pandas as pd
import numpy as np
import xlrd
import unicodedata
import openpyxl
from sklearn.preprocessing import StandardScaler  # 导入标准化数据
from sklearn.metrics import mean_squared_error  # 均方误差

###################以下模型，按需导入
from sklearn.linear_model import Ridge


def is_number(s):
    try:
        if s=='' or s=='nan' or '_' in s or s[-1]=='.':
            return False
        float(s)
        return True
    except ValueError:
        pass
    try:
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False

def whether_list(data):  # 判断是否列表是就返回第一元素
    if isinstance(data, np.ndarray) or isinstance(data, list):
        return data[0]
    else:
        return data

def return_number(s,reserved_decimal='off'):
    if not is_number(s) or s == 'inf' or s=='-inf':
        return s
    else:
        if ('.' in s or 'e' in s )and set(s.split('.')[-1])!={'0'}:
            if reserved_decimal!='off':
                return round(float(s),reserved_decimal)
            return float(s)
        else:
            if '.' in s and set(s.split('.')[-1])=={'0'}:
                s=s[:len(s)-len(s.split('.')[-1])-1]
            try:
                return int(s)
            except:
                return s

def read_xlsx(name):
    workbook = xlrd.open_workbook(name)  # 打开工作簿
    data_sheet = workbook.sheets()[0]  # 通过索引获取
    rowNum = data_sheet.nrows  # sheet行数
    # 获取所有单元格的内容
    data_train = []
    for i in range(1,rowNum):#不要标题行
        data_train.append(data_sheet.row_values(i))

    x_train = []
    y_train = []
    for i in range(len(data_train)):
        for j in range(len(data_train[i])):
            if j == 0:
                try:
                    y_train.append(float(data_train[i][j]))
                except:
                    y_train.append(data_train[i][j])
            else:
                try:
                    x_train.append(float(data_train[i][j]))
                except:
                    x_train.append(data_train[i][j])
    x_train = np.array(x_train)
    x_train = x_train.reshape(-1, len(data_train[0])-1)
    y_train = np.array(y_train)
    y_train = y_train.reshape(-1, 1)

    data_sheet = workbook.sheets()[1]  # 通过索引获取
    rowNum = data_sheet.nrows  # sheet行数
    # 获取所有单元格的内容
    data_test = []
    for i in range(1,rowNum):#不要标题行
        data_test.append(data_sheet.row_values(i))
    x_test = []
    y_test = []
    for i in range(len(data_test)):
        for j in range(len(data_test[i])):
            if j == 0:
                try:
                    y_test.append(float(data_test[i][j]))
                except:
                    y_test.append(data_test[i][j])
            else:
                try:
                    x_test.append(float(data_test[i][j]))
                except:
                    x_test.append(data_test[i][j])
    x_test = np.array(x_test)
    x_test = x_test.reshape(-1, len(data_test[0])-1)
    y_test = np.array(y_test)
    y_test = y_test.reshape(-1, 1)

    x_all=np.concatenate((x_train, x_test), axis=0)
    y_all=np.concatenate((y_train, y_test), axis=0)
    #导出
    result={}
    result['x_all']=x_all
    result['y_all']=y_all
    result['x_train']=x_train
    result['y_train']=y_train
    result['x_test']=x_test
    result['y_test']=y_test
    return result

def write_xlsx(name, datalist= [[]],turn_float=True):
    # 设置文件 mingc
    workbook = openpyxl.Workbook()
    if isinstance(datalist,pd.core.frame.DataFrame):
        datalist=np.array(datalist).tolist()##panda转list
    # # 创建一张新表
    sheet = workbook.create_sheet('demo',index=0)
    for row in range(len(datalist)):
        for column in range(len(datalist[row])):
            if turn_float:
                tem=return_number(str(datalist[row][column]))
            else:
                tem=str(datalist[row][column])
            sheet.cell(row = row+1, column = column+1).value = tem
    if not os.path.exists(os.path.dirname(name)):# 如果不存在则创建目录
        os.makedirs(os.path.dirname(name))
    workbook.save(name)
    print("创建" + name + '成功')

# 计算RRMSE
def RRMSE(y_true, y_clf):#(%)
    '''计算RRMSE(预测值,真实值)，注意真实值得和不能为0，注意真实值均为正数'''
    sum = 0
    for i in range(len(y_true)):
        sum = sum + y_true[i]
        sum = abs(sum)
    return pow(mean_squared_error(y_true, y_clf), 0.5) * len(y_true) / sum

# 计算AUC
def Auc(y_true, y_clf):
    try:
        if isinstance(y_true[0],np.ndarray):
            y_true=y_true.reshape(-1).tolist()
    except:
        pass
    try:
        if isinstance(y_clf[0],np.ndarray):
            y_clf=y_clf.reshape(-1).tolist()
    except:
        pass
    data_dict={
        'y_true':y_true,
        'y_clf': y_clf
    }
    df=pd.DataFrame(data_dict)
    df['rank'] = df['y_clf'].rank()

    condition = df['y_true'] > 0
    score1 = df[condition]['rank'].sum()
    score2=score1-sum(condition)*(sum(condition)+1)/2
    result=score2/sum(condition)/(len(df)-sum(condition))
    return result

def Calculate_Standard(Standard, y_true, y_clf):
    result = None
    if Standard == 'RRMSE':
        result = RRMSE(y_true, y_clf)
    elif Standard == 'Auc':
        result = Auc(y_true, y_clf)  # 变成负数为了统一越小越好

    if isinstance(result, np.ndarray) or isinstance(result, list):
        if len(result) == 1:
            result = whether_list(result)
    return result

def inversion_xlsx(filename_clf,filename_xlsx,Standard):
    '''根据拟合器反演xlsx数据,没有行'''
    save = shelve.open(filename_clf[:-4])
    clf = save['clf']
    ss = save['ss']

    data=read_xlsx(filename_xlsx)
    data1=copy.deepcopy(data['x_test'])
    data1=ss.transform(data1)
    y_clf=clf.predict(data1)
    y_clf = y_clf.reshape(-1, 1)
    score = Calculate_Standard(Standard, data['y_test'], y_clf)
    print(Standard,':',score)
    result=np.concatenate((y_clf, data['x_test']), axis=1)

    title = [0]
    for i in range(len(data['x_test'][0])):
        title.append(i)
    title= np.array(title).reshape(-1,len(title))
    result=np.concatenate((title, result), axis=0)
    filename=filename_xlsx[:-5]+'_inv.xlsx'
    write_xlsx(filename,result)

def run_one_n(filename='', Type='', par='', Type_par='', Standard='RRMSE'):
    # 导入数据
    data=read_xlsx(filename)
    data1=copy.deepcopy(data)
    # 预处理
    ss = StandardScaler()
    data1['x_train'] = ss.fit_transform(data1['x_train'])
    data1['x_test'] = ss.transform(data1['x_test'])
    # 运行
    if not Type_par:
        clf = eval(Type + '(' + par + ')')  # 根据类型及其参数建模
    else:
        clf = eval(Type_par)
    clf.fit(data1['x_train'], data1['y_train'])
    y_clf = clf.predict(data1['x_test'])
    y_clf = y_clf.reshape(-1, 1)
    # 输出
    score = Calculate_Standard(Standard, data1['y_test'], y_clf)
    print(Standard,':',score)
    save_name=filename[:-5]+Type
    save = shelve.open(save_name)
    save['clf'] = clf
    save['ss'] = ss
    save.close()
    result=np.concatenate((y_clf, data['x_test']), axis=1)
    title = [0]
    for i in range(len(data['x_train'][0])):
        title.append(i)
    title= np.array(title).reshape(-1,len(title))
    result=np.concatenate((title, result), axis=0)
    filename=filename_xlsx[:-5]+'_inv.xlsx'
    write_xlsx(filename,result)
    return score

if __name__ == '__main__':
    ##创建模型并离线模型
    # filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y_Ridge.xlsx'
    # run_one_n(filename_xlsx,Type_par="Ridge(random_state=8)",Standard='Auc')
    #
    # filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y1_Ridge.xlsx'
    # run_one_n(filename_xlsx,Type_par="Ridge(random_state=8)",Standard='RRMSE')
    #
    # filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y2_Ridge.xlsx'
    # run_one_n(filename_xlsx,Type_par="Ridge(random_state=8)",Standard='RRMSE')
    #
    # filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y3_Ridge.xlsx'
    # run_one_n(filename_xlsx,Type_par="Ridge(random_state=8)",Standard='RRMSE')

    # ##调用离线模型
    filename_clf=r'D:\atest\lize\260610创建并离线模型\y_Ridge.dat'
    filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y_Ridge.xlsx'
    inversion_xlsx(filename_clf,filename_xlsx,Standard='Auc')

    filename_clf=r'D:\atest\lize\260610创建并离线模型\y1_Ridge.dat'
    filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y1_Ridge.xlsx'
    inversion_xlsx(filename_clf,filename_xlsx,Standard='RRMSE')

    filename_clf=r'D:\atest\lize\260610创建并离线模型\y2_Ridge.dat'
    filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y2_Ridge.xlsx'
    inversion_xlsx(filename_clf,filename_xlsx,Standard='RRMSE')

    filename_clf=r'D:\atest\lize\260610创建并离线模型\y2_Ridge.dat'
    filename_xlsx=r'D:\atest\lize\260610创建并离线模型\y2_Ridge.xlsx'
    inversion_xlsx(filename_clf,filename_xlsx,Standard='RRMSE')